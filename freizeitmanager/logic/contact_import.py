"""Import einer Personenliste aus CSV oder Excel.

Der Zweck ist der Erststart: wer seine Kontakte schon in einer Tabelle hat,
soll sie nicht abtippen muessen. Die Spalten werden erkannt, statt eine feste
Reihenfolge zu verlangen - echte Adressbuchexporte haben nie dieselbe.

Bewusste Entscheidungen:

* **Lesen und Schreiben sind getrennt.** ``read_table`` und ``build_preview``
  fassen die Datenbank nicht an. Erst ``apply_preview`` schreibt, und nur die
  Zeilen, die der Nutzer freigegeben hat.
* **Nichts wird ungefragt ueberschrieben.** Ein Name, den es schon gibt, wird
  als Konflikt gemeldet. Standard ist Ueberspringen.
* **Eine unlesbare Zeile kippt nicht den Import.** Sie wird mit Grund gemeldet
  und uebersprungen.
"""
from __future__ import annotations

import csv
import io
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from enum import Enum
from pathlib import Path
from typing import Any, Iterable, Sequence

CSV_SUFFIXES = {".csv", ".txt", ".tsv"}
EXCEL_SUFFIXES = {".xlsx", ".xlsm"}

# Ein Jahr, das "Jahrgang unbekannt" bedeutet. Viele Adressbuecher exportieren
# Tag und Monat ohne Jahr; ein erfundener Jahrgang waere eine Falschangabe.
YEAR_UNKNOWN = 1900

# Befunde einer Zeile. Bewusst Schluessel statt Text: sie werden in der
# Vorschau angezeigt und muessen der gewaehlten Sprache folgen.
PROBLEM_NO_NAME = "import.problem_no_name"
PROBLEM_BAD_DATE = "import.problem_bad_date"
PROBLEM_FUTURE_DATE = "import.problem_future_date"


class ImportError_(ValueError):
    """Die Datei laesst sich nicht als Tabelle lesen.

    Traegt einen Uebersetzungsschluessel statt eines fertigen Satzes: die
    Meldung landet in der Oberflaeche und muss der gewaehlten Sprache folgen.
    Der Ausnahmetext bleibt der Schluessel - der ist fuers Log eindeutig.
    """

    def __init__(self, key: str, **params: Any):
        self.key = key
        self.params = params
        super().__init__(key)


ERROR_UNSUPPORTED = "import.error_unsupported"
ERROR_ENCODING = "import.error_encoding"
ERROR_EMPTY = "import.error_empty"
ERROR_NO_ROWS = "import.error_no_rows"
ERROR_NEEDS_OPENPYXL = "import.error_needs_openpyxl"
ERROR_EXCEL_UNREADABLE = "import.error_excel_unreadable"
ERROR_SHEET_EMPTY = "import.error_sheet_empty"


class RowAction(str, Enum):
    CREATE = "create"      # Neuer Kontakt.
    FILL = "fill"          # Bestehender Kontakt, nur leere Felder auffuellen.
    SKIP = "skip"          # Nichts tun.


# Spaltenerkennung. Links die Zielspalte, rechts die Ueberschriften, die dafuer
# durchgehen - deutsche und englische Adressbuchexporte gemischt.
COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "name": ("name", "vollername", "fullname", "displayname", "kontakt", "person",
             "anzeigename", "contact", "bezeichnung"),
    "first_name": ("vorname", "firstname", "givenname", "rufname"),
    "last_name": ("nachname", "name2", "lastname", "familyname", "surname",
                  "familienname", "zuname"),
    "birthday": ("geburtstag", "geburtsdatum", "geburt", "birthday", "birthdate",
                 "dateofbirth", "dob", "gebdatum", "gebtag"),
    "notes": ("notiz", "notizen", "bemerkung", "bemerkungen", "note", "notes",
              "kommentar", "comment", "beschreibung"),
    "group": ("gruppe", "group", "kategorie", "category", "kreis"),
    "email": ("email", "emailadresse", "mail", "epost"),
    "phone": ("telefon", "telefonnummer", "phone", "mobil", "handy", "mobile",
              "tel", "rufnummer"),
}

# Datumsformate in der Reihenfolge, in der sie probiert werden. Deutsch zuerst,
# weil die Oberflaeche deutsch ist: 03.04. ist hier der 3. April, nicht der
# 4. Maerz. ISO steht davor, weil es eindeutig ist.
_DATE_FORMATS_WITH_YEAR = (
    "%Y-%m-%d", "%Y/%m/%d",
    "%d.%m.%Y", "%d.%m.%y",
    "%d/%m/%Y", "%d-%m-%Y",
    "%d. %B %Y", "%d. %b %Y",
)
_DATE_FORMATS_WITHOUT_YEAR = ("%d.%m.", "%d.%m", "%m-%d", "%d/%m")
# strptime ohne Jahresangabe ist mehrdeutig und wird ab Python 3.15
# nicht mehr unterstuetzt; deshalb haengen wir das Ersatzjahr selbst an.


def _normalize_header(value: Any) -> str:
    """Ueberschrift auf einen vergleichbaren Kern reduzieren."""
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("ß", "ss").lower()
    return re.sub(r"[^a-z0-9]", "", text)


def detect_columns(headers: Sequence[Any]) -> dict[str, int]:
    """Ordnet Zielfeldern die Spaltennummer zu. Erste Fundstelle gewinnt."""
    found: dict[str, int] = {}
    normalized = [_normalize_header(h) for h in headers]
    for field_name, aliases in COLUMN_ALIASES.items():
        for index, header in enumerate(normalized):
            if not header:
                continue
            if header in aliases and field_name not in found:
                found[field_name] = index
                break
    # Zweiter Durchgang: Teiltreffer, damit "Geburtstag (TT.MM.)" auch zaehlt.
    for field_name, aliases in COLUMN_ALIASES.items():
        if field_name in found:
            continue
        for index, header in enumerate(normalized):
            if header and index not in found.values() and any(a in header for a in aliases):
                found[field_name] = index
                break
    return found


def parse_birthday(value: Any) -> tuple[date | None, bool, str]:
    """Geburtstag lesen.

    Rueckgabe: (Datum, Jahrgang bekannt, Befundschluessel). Der Schluessel ist
    leer, wenn nichts zu melden ist. Ohne Jahrgang wird ``YEAR_UNKNOWN``
    gesetzt, damit Tag und Monat nicht verloren gehen.
    """
    if value is None:
        return None, True, ""
    if isinstance(value, datetime):
        return value.date(), True, ""
    if isinstance(value, date):
        return value, True, ""
    text = str(value).strip()
    if not text:
        return None, True, ""
    for fmt in _DATE_FORMATS_WITH_YEAR:
        try:
            parsed = datetime.strptime(text, fmt).date()
        except ValueError:
            continue
        if parsed > date.today():
            # Zweistellige Jahre laufen sonst in die Zukunft: '65 -> 2065.
            try:
                parsed = parsed.replace(year=parsed.year - 100)
            except ValueError:
                return None, True, PROBLEM_FUTURE_DATE
        return parsed, True, ""
    for fmt in _DATE_FORMATS_WITHOUT_YEAR:
        try:
            parsed = datetime.strptime(f"{text} {YEAR_UNKNOWN}", f"{fmt} %Y").date()
        except ValueError:
            continue
        return date(YEAR_UNKNOWN, parsed.month, parsed.day), False, ""
    return None, True, PROBLEM_BAD_DATE


@dataclass
class ImportRow:
    """Eine Zeile der Datei, fertig gelesen und bewertet."""
    line: int
    name: str = ""
    birthday: date | None = None
    birthday_has_year: bool = True
    raw_birthday: str = ""   # Rohtext, damit ein unlesbares Datum zeigbar bleibt.
    notes: str = ""
    group: str = ""
    email: str = ""
    phone: str = ""
    action: RowAction = RowAction.CREATE
    existing_id: int | None = None
    problems: list[str] = field(default_factory=list)

    @property
    def is_conflict(self) -> bool:
        return self.existing_id is not None

    @property
    def is_usable(self) -> bool:
        return bool(self.name) and PROBLEM_NO_NAME not in self.problems


@dataclass
class ImportPreview:
    """Was der Import tun wuerde - vor jedem Schreibzugriff."""
    rows: list[ImportRow] = field(default_factory=list)
    detected: dict[str, int] = field(default_factory=dict)
    headers: list[str] = field(default_factory=list)

    @property
    def new_rows(self) -> list[ImportRow]:
        return [r for r in self.rows if r.is_usable and not r.is_conflict]

    @property
    def conflicts(self) -> list[ImportRow]:
        return [r for r in self.rows if r.is_usable and r.is_conflict]

    @property
    def unusable(self) -> list[ImportRow]:
        return [r for r in self.rows if not r.is_usable]

    @property
    def with_birthday(self) -> int:
        return sum(1 for r in self.rows if r.is_usable and r.birthday is not None)


def _sniff_dialect(sample: str) -> csv.Dialect | type[csv.Dialect]:
    try:
        return csv.Sniffer().sniff(sample, delimiters=";,\t|")
    except csv.Error:
        # Semikolon ist der Normalfall bei deutschem Excel-CSV.
        return csv.excel_tab if "\t" in sample else csv.excel


def read_table(path: Path | str) -> tuple[list[str], list[list[Any]]]:
    """Liest CSV oder Excel als Kopfzeile plus Datenzeilen."""
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix in EXCEL_SUFFIXES:
        return _read_excel(path)
    if suffix in CSV_SUFFIXES:
        return _read_csv(path)
    raise ImportError_(ERROR_UNSUPPORTED, suffix=path.suffix or path.name)


def _read_csv(path: Path) -> tuple[list[str], list[list[Any]]]:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:  # pragma: no cover - latin-1 nimmt jedes Byte an
        raise ImportError_(ERROR_ENCODING)
    if not text.strip():
        raise ImportError_(ERROR_EMPTY)
    reader = csv.reader(io.StringIO(text), dialect=_sniff_dialect(text[:4096]))
    rows = [row for row in reader if any(str(cell).strip() for cell in row)]
    if not rows:
        raise ImportError_(ERROR_NO_ROWS)
    return [str(c).strip() for c in rows[0]], [list(r) for r in rows[1:]]


def _read_excel(path: Path) -> tuple[list[str], list[list[Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - Abhaengigkeit ist gepinnt
        raise ImportError_(ERROR_NEEDS_OPENPYXL) from exc
    try:
        book = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ImportError_(ERROR_EXCEL_UNREADABLE, detail=str(exc)) from exc
    try:
        sheet = book.active
        rows = [list(r) for r in sheet.iter_rows(values_only=True)
                if any(str(c).strip() for c in r if c is not None)]
    finally:
        book.close()
    if not rows:
        raise ImportError_(ERROR_SHEET_EMPTY)
    return [str(c or "").strip() for c in rows[0]], rows[1:]


def _cell(row: Sequence[Any], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    value = row[index]
    return "" if value is None else str(value).strip()


def build_preview(
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    existing: dict[str, int] | None = None,
) -> ImportPreview:
    """Bewertet die Zeilen, ohne etwas zu schreiben.

    ``existing`` bildet den kleingeschriebenen Namen auf die Kontakt-ID ab.
    """
    detected = detect_columns(headers)
    known = {k.strip().lower(): v for k, v in (existing or {}).items()}
    preview = ImportPreview(detected=detected, headers=[str(h) for h in headers])

    for offset, raw in enumerate(rows, start=2):
        item = ImportRow(line=offset)
        name = _cell(raw, detected.get("name"))
        if not name:
            first = _cell(raw, detected.get("first_name"))
            last = _cell(raw, detected.get("last_name"))
            name = " ".join(part for part in (first, last) if part)
        item.name = name
        if not name:
            item.problems.append(PROBLEM_NO_NAME)
            item.action = RowAction.SKIP
            preview.rows.append(item)
            continue

        birthday_raw = raw[detected["birthday"]] if "birthday" in detected and detected["birthday"] < len(raw) else None
        item.raw_birthday = "" if birthday_raw is None else str(birthday_raw).strip()
        item.birthday, item.birthday_has_year, problem = parse_birthday(birthday_raw)
        if problem:
            item.problems.append(problem)

        item.notes = _cell(raw, detected.get("notes"))
        item.group = _cell(raw, detected.get("group"))
        item.email = _cell(raw, detected.get("email"))
        item.phone = _cell(raw, detected.get("phone"))

        existing_id = known.get(name.strip().lower())
        if existing_id is not None:
            item.existing_id = existing_id
            # Bestehende Kontakte bleiben unangetastet, bis der Nutzer
            # ausdruecklich etwas anderes waehlt.
            item.action = RowAction.SKIP
        preview.rows.append(item)
    return preview


@dataclass
class ImportResult:
    """Was tatsaechlich geschrieben wurde."""
    created: int = 0
    filled: int = 0
    skipped: int = 0
    failed: list[str] = field(default_factory=list)

    @property
    def total_written(self) -> int:
        return self.created + self.filled


def existing_names(session) -> dict[str, int]:
    """Vorhandene Kontaktnamen kleingeschrieben auf ihre ID."""
    from sqlalchemy import select

    from freizeitmanager.database.models import Contact

    return {
        str(name).strip().lower(): contact_id
        for contact_id, name in session.execute(select(Contact.id, Contact.name))
    }


def apply_preview(session, preview: ImportPreview) -> ImportResult:
    """Schreibt genau die Zeilen, die auf CREATE oder FILL stehen."""
    from freizeitmanager.database.models import Contact
    from freizeitmanager.logic.contact_service import create_contact

    result = ImportResult()
    for row in preview.rows:
        if row.action == RowAction.SKIP or not row.is_usable:
            result.skipped += 1
            continue
        try:
            if row.action == RowAction.CREATE:
                create_contact(
                    session,
                    row.name,
                    groups=[row.group] if row.group else None,
                    notes=row.notes or None,
                    birthday=row.birthday,
                    birthday_has_year=row.birthday_has_year,
                )
                result.created += 1
            else:  # FILL - bestehende Werte bleiben stehen.
                contact = session.get(Contact, row.existing_id)
                if contact is None:
                    result.failed.append(f"Zeile {row.line}: Kontakt nicht mehr vorhanden")
                    continue
                changed = False
                if contact.birthday is None and row.birthday is not None:
                    contact.birthday = row.birthday
                    contact.birthday_has_year = row.birthday_has_year
                    changed = True
                if not contact.notes and row.notes:
                    contact.notes = row.notes
                    changed = True
                result.filled += 1 if changed else 0
                result.skipped += 0 if changed else 1
        except Exception as exc:  # Eine kaputte Zeile kippt nicht den Rest.
            session.rollback()
            result.failed.append(f"Zeile {row.line} ({row.name}): {exc}")
    session.flush()
    return result
